from ResultApp.models import Aset, Student
import openpyxl
import uuid

def generatelist(excel_file,asetid):
    print("Starting............")
    wb = openpyxl.Workbook()        
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["Students"]
    row_start=2
    total = ws.max_row
    studentlist=[]
    existinglist=[]
    myaset= Aset.objects.get(asetid=asetid)
    try :

            for i in range(row_start,total+ row_start):
                  
                  matricno = ws.cell(i, 1).value
                  print(i,matricno)
                  if matricno is not None:                         
                        #    created= Student.objects.create(    
                        #         matricno = matricno,  
                        #         studentGuId  =  uuid.uuid4(),
                        #         asetid = myaset,    
                        asession= ws.cell(i, 2).value
                        faculty=  ws.cell(i, 3).value    
                        department=  ws.cell(i, 4).value
                        programme=  ws.cell(i, 5).value  
                        option = ws.cell(i, 6).value 

                        applicationnumber=  ws.cell(i, 7).value
                        formno=  ws.cell(i,8).value     
                        surname=  ws.cell(i, 9).value  
                        middlename=  ws.cell(i, 10).value 
                        firstname= ws.cell(i, 11).value
                        presentstate=  ws.cell(i, 12).value
                        emailaddress= ws.cell(i, 13).value
                        phonenumber=  ws.cell(i, 14).value
                        # #student,graduate,suspended,leave,withdraw
                        #         graduatingsession=''})    
                        try :
                         obj,created= Student.objects.get_or_create(    
                                matricno = matricno, 
                                defaults={ 
                                'studentGuId'  :  uuid.uuid4(),
                                'aset' : myaset,    
                                'asession': asession,   
                                'faculty': faculty,    
                                'department': department,
                                'programme': programme,  
                                'option':option,
                                'applicationnumber': applicationnumber,
                                'formno': formno,     
                                'surname': surname,  
                                'middlename': middlename, 
                                'firstname':firstname,
                                'presentstate': presentstate,#student,graduate,suspended,leave,withdraw
                                'emailaddress':emailaddress,
                                'phonenumber':phonenumber,
                                'password':matricno,
                                'graduatingsession':''})  
                        except Exception as inst:
                            print("I  raise error on => ", matricno)
                            print(inst) 
                         


                        if created :
                            studentlist.append(obj)
                        else:
                            existinglist.append(obj)
                        #ns.save()
                        print(obj)
                        
    except  Exception as inst:
          print("I  raise error on => ", matricno)
          print(inst)   
                         
    return studentlist, existinglist
            
